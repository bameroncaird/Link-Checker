from test_link import test_link
import openpyxl
import os

LINK_COL = 'C'
CODE_COL = 'K'
CODE_DESC_COL = 'L'
CODE_LINK_COL = 'M'
SECOND_LINK_COL = 'N'
EXCEL_FILE_NAME = 'website_links.xlsx'

def main():
    wb, excel_file = initialize_excel_file()
    last_row = excel_file.max_row
    first_row = find_start(excel_file)
    current = first_row

    print(first_row)

    while True:
        if first_row > last_row:
            break
        batch_size = get_batch_size()
        if first_row + batch_size > last_row:
            batch_size = last_row - first_row

        for _ in range(batch_size):
            link = excel_file[LINK_COL + str(current)].value
            result = test_link(link)
            print(link, result)
            excel_file[CODE_COL + str(current)].value = result[0]
            excel_file[CODE_DESC_COL + str(current)].value = result[1]
            excel_file[CODE_LINK_COL + str(current)].value = result[2]
            excel_file[SECOND_LINK_COL + str(current)].value = link
            wb.save(EXCEL_FILE_NAME)
            current += 1

        open_excel_file()
        print("Remember to save and close the file.")
        choice = get_continuation_choice()
        if choice == 'N' or choice == 'n':
            break

    print("Done for now.")


# Returns the excel file.
def initialize_excel_file():
    wb = openpyxl.load_workbook(EXCEL_FILE_NAME)
    sheet = wb.sheetnames[0]
    excel_sheet = wb[sheet]
    return wb, excel_sheet

# Returns the first empty row in the response code column.
def find_start(excel_sheet):
    row = 1
    for cell in excel_sheet[CODE_COL]:
        if cell.value is None:
            break
        row += 1
    return row

# Gets the user's input for batch size and doesn't allow them to mess up.
def get_batch_size():
    size = None
    while size is None:
        try:
            size = int(input("Enter the batch size: "))
            if size > 0:
                return size
            else:
                size = None
                print("You did not enter a valid batch size and will have to try again.")
        except:
            print("You did not enter a valid batch size and will have to try again.")

# Asks the user if they would like to continue a loop.
def get_continuation_choice():
    continuation_set = {'Y', 'y', 'N', 'n'}
    choice = str(input("More batches? Y/N: "))
    while choice not in continuation_set:
        print("You did not enter a valid choice and will have to try again.")
        choice = str(input("More batches? Y/N: "))
    return choice

# Opens the Excel file so the user can see the results.
def open_excel_file():
    os.system("open " + EXCEL_FILE_NAME)

def rerun():
    wb, excel_file = initialize_excel_file()
    last_row = excel_file.max_row
    current_row = 17793
    while current_row <= last_row:
        old_code = int(excel_file[CODE_COL + str(current_row)].value)
        if old_code != 200:
            link = excel_file[LINK_COL + str(current_row)].value
            new_info = test_link(link)
            new_code = int(new_info[0])
            if new_code != old_code:
                excel_file[CODE_COL + str(current_row)].value = new_code
                excel_file[CODE_DESC_COL + str(current_row)].value = new_info[1]
                excel_file[CODE_LINK_COL + str(current_row)].value = new_info[2]
                excel_file[SECOND_LINK_COL + str(current_row)].value = link
                wb.save(EXCEL_FILE_NAME)
                print("Updated: row " + str(current_row) + "/" + str(last_row))
        current_row += 1
    open_excel_file()
    print("Remember to save and close the file.")

# main()
rerun()